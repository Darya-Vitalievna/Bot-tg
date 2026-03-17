from pathlib import Path
from tempfile import NamedTemporaryFile
from typing import Dict, Iterable, List, Tuple

from . import db_help as db


GUILD_TITLES: Dict[int, str] = {
    1: "Блоггинг",
    2: "Карьера",
    3: "Личная эффективность",
    4: "Спорт",
}


def _guild_name(guild_id: int) -> str:
    return GUILD_TITLES.get(int(guild_id or 0), f"Гильдия {guild_id}")


def _safe_username(username: str) -> str:
    if not username:
        return ""
    value = str(username).strip().lstrip("@")
    return f"@{value}" if value else ""


def _safe_text(value) -> str:
    if value is None:
        return ""
    return str(value)


def _autofit_columns(ws) -> None:
    for column_cells in ws.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            cell_value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(cell_value))
        ws.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 50)


def _style_sheet(ws, title: str) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = True

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.row_dimensions[1].height = 24
    ws.title = title
    _autofit_columns(ws)


def _build_workbook(sheet_rows: Iterable[Tuple[str, List[str], List[List[str]]]]) -> Path:
    from openpyxl import Workbook

    workbook = Workbook()
    first_sheet = True

    for title, headers, rows in sheet_rows:
        if first_sheet:
            ws = workbook.active
            first_sheet = False
        else:
            ws = workbook.create_sheet()

        ws.append(headers)
        for row in rows:
            ws.append(row)
        _style_sheet(ws, title)

    temp_file = NamedTemporaryFile(prefix="bookbot_export_", suffix=".xlsx", delete=False)
    temp_path = Path(temp_file.name)
    temp_file.close()
    workbook.save(temp_path)
    workbook.close()
    return temp_path


def build_books_exercises_export() -> Path:
    exercise_rows, book_rows = db.export_books_and_exercises_rows(month_key=db.current_month_key())

    combined_headers = [
        "Тип",
        "Дата",
        "Юзер ID",
        "Telegram",
        "Имя",
        "Упражнение",
        "Количество",
        "Единица",
        "Принято ИИ",
        "Книга",
        "Результат",
        "Вопросы",
        "Ответы",
        "Video File ID",
    ]
    combined_rows = [
        [
            "Упражнение",
            _safe_text(row.get("created_at")),
            _safe_text(row.get("user_id")),
            _safe_username(row.get("username")),
            _safe_text(row.get("full_name")),
            _safe_text(row.get("exercise_name")),
            _safe_text(row.get("amount")),
            _safe_text(row.get("unit")),
            _safe_text(row.get("ai_status") or "Принято"),
            "",
            "",
            "",
            "",
            _safe_text(row.get("video_file_id")),
        ]
        for row in exercise_rows
    ]

    combined_rows.extend(
        [
            [
                "Книга",
                _safe_text(row.get("created_at")),
                _safe_text(row.get("user_id")),
                _safe_username(row.get("username")),
                _safe_text(row.get("full_name")),
                "",
                "",
                "",
                "",
                _safe_text(row.get("book_name")),
                _safe_text(row.get("result_text")),
                _safe_text(row.get("questions_text")),
                _safe_text(row.get("answers_text")),
                "",
            ]
            for row in book_rows
        ]
    )

    return _build_workbook(
        [
            ("Книги и упражнения", combined_headers, combined_rows),
        ]
    )


def build_help_export() -> Path:
    request_rows, answer_rows = db.export_help_requests_and_answers_rows(month_key=db.current_month_key())
    request_map = {int(row["request_id"]): row for row in request_rows if row.get("request_id") is not None}

    help_headers = [
        "Дата ответа",
        "Answer ID",
        "Request ID",
        "Эксперт Telegram",
        "Эксперт имя",
        "Пользователь Telegram",
        "Пользователь имя",
        "Суть запроса",
        "Желаемый результат",
        "Ответ эксперта",
        "Гильдия",
        "Эксперт ID",
        "User ID",
    ]
    help_sheet_rows = [
        [
            _safe_text(row.get("created_at")),
            _safe_text(row.get("answer_id")),
            _safe_text(row.get("request_id")),
            _safe_username(row.get("expert_username")),
            _safe_text(row.get("expert_full_name")),
            _safe_username(row.get("student_username")),
            _safe_text(row.get("student_full_name")),
            _safe_text(row.get("problem_essence") or request_map.get(int(row.get("request_id") or 0), {}).get("problem_essence")),
            _safe_text(row.get("desired_result") or request_map.get(int(row.get("request_id") or 0), {}).get("desired_result")),
            _safe_text(row.get("answer_text")),
            _guild_name(row.get("guild_id") or request_map.get(int(row.get("request_id") or 0), {}).get("guild_id")),
            _safe_text(row.get("expert_user_id")),
            _safe_text(row.get("student_user_id") or request_map.get(int(row.get("request_id") or 0), {}).get("student_user_id")),
        ]
        for row in answer_rows
    ]

    return _build_workbook(
        [
            ("Запросы и ответы", help_headers, help_sheet_rows),
        ]
    )


def build_experts_export(experts: List[Dict[str, str]]) -> Path:
    headers = [
        "№",
        "Эксперт Telegram",
        "Эксперт имя",
        "Гильдии",
        "Эксперт ID",
    ]
    rows = [
        [
            _safe_text(expert.get("number")),
            _safe_username(expert.get("username")),
            _safe_text(expert.get("full_name")),
            _safe_text(expert.get("guilds")),
            _safe_text(expert.get("user_id")),
        ]
        for expert in experts
    ]

    return _build_workbook(
        [
            ("Список экспертов", headers, rows),
        ]
    )
